from openpyxl import Workbook, load_workbook

wb = load_workbook("colors.xlsx")
ws = wb.active

i = 0
while (i <= 255):
	ws["A" + str(i + 1)].value = str(hex(i))
	i += 1

wb.save("colors.xlsx")